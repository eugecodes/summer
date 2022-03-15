from __future__ import unicode_literals
import os

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "exportAbroad.settings")

import django
django.setup()
import xlrd
from openpyxl import load_workbook
from newmarkets.models import HSProduct, Country, IntracenImportData

PATH = '/home/hakim/Dropbox/Data5'
def get_hs_and_description(product):
	product = product.split()
	hs_number = product[1]
	description = "".join(product[2:])
	return hs_number, description


for dirName, subdirList, fileList in os.walk(PATH):
    #print('Found directory: %s' % dirName)
    for fname in fileList:
    	if fname.endswith(".xlsx"):
        	#work on the file with openpyxl
        	wb = load_workbook(filename = os.path.join(dirName,fname))
        	ws = wb.active
        	prod = ws.cell('A2').value
        	print get_hs_and_description(prod)



def parse_file(datafile):
	pass
def populate():
	pass


if __name__ == '__main__':
	main()